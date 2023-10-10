# -*- coding: utf-8 -*-
"""
Created on Sun Sep  3 22:17:14 2023

@author: Johnson
"""

import requests as req
import openpyxl as xl

# max_page,useragent請依個人狀況設定
max_page=43
useragent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Safari/537.36'

# programing start
try:
    wb=xl.Workbook()
    ws=wb.active
    title=['課名','作者','價格','預售價','販售數']
    ws.append(title)
    
    setting={
        'User-Agent':useragent
        }
    
    for number in range(max_page+1):
        # url is iterator
        url='https://api.hahow.in/api/products/search?category=COURSE&limit=24&page={}&sort=TRENDING'.format(number)
        print(url)
        rsp=req.get(url,headers=setting)
        # 查看response狀態
        print(rsp.status_code)
        root_json=rsp.json()
        
        for data in root_json['data']['courseData']['products']:
            course=[]
            course.append(data['title'])
            course.append(data['owner']['name'])
            course.append(data['price'])
            course.append(data['preOrderedPrice'])
            course.append(data['numSoldTickets'])
            ws.append(course)
            
    wb.save('howhow_course.xlsx')
except Exception as e:
    print(e)
    
finally:
    print('Done for scrapy')

#%% # 上/下為互補程式碼，請分段執行

import openpyxl as xl

try:
    wb=xl.load_workbook('howhow_course.xlsx')
    
    ws=wb.active
    
    fontstyle=xl.styles.Font(name='標楷體',bold=True,color='1e90ff')
    
    for col in range(1,6):
        char=xl.utils.get_column_letter(col)
        ws[char+str(1)].font=fontstyle
        
    wb.save('howhow_course.xlsx')
except Exception as e:
    print(e)
finally:
    print('Done for modify fonts')